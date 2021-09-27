import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
from scipy import sparse
from scipy.sparse.linalg import spsolve
from scipy.spatial import ConvexHull
import openpyxl as xl
import sys
import tkinter as tk
from tkinter import filedialog

class Vibration:
    def __init__(self, name):
        self.name = name
        self.frequence = []

class IR_DATA:

    def __init__(self, type, wb_name, sheetname):
        self.type = type
        self.sheetname = sheetname
        self.wb_name = wb_name
        self.wb = None
        self.sheet = None
        self.rows = None
        self.vibrations = []
        self.WorkBook()
        self.readSheet()
        self.getDim()
        self.getWN()

    def WorkBook(self):
        wb = xl.load_workbook(self.wb_name, data_only = True)
        self.wb = wb

    def readSheet(self):
        self.sheet = self.wb[self.sheetname]

    def getDim(self):
        sheet = self.sheet
        rowNum = 0
        i = 2
        while True:
            if sheet.cell(row=i, column = 1).value is not None:
                rowNum += 1
            else:
                break
            i+=1
        self.rows = rowNum
        
    def getWN(self):
        prevType = self.sheet.cell(row = 2, column = 1).value
        actVibration = Vibration(prevType)
        i = 2
        while i <= self.rows:
            actType = self.sheet.cell(row = i, column = 1).value
            if actType != prevType:
                self.vibrations.append(actVibration)
                prevType = actType
                actVibration = Vibration(actType)
 
            wnMax = int(self.sheet.cell(row = i, column = 2).value)
            wnMin = int(self.sheet.cell(row = i, column = 3).value)
            intensity = self.sheet.cell(row = i, column = 4).value
            remark = self.sheet.cell(row = i, column = 5).value

            data = (wnMax, wnMin, intensity, remark)
            actVibration.frequence.append(data)

            i+=1

def Transpone(y):

    new_y = []

    for data in y:
        data = round(100 - data,4)
        new_y.append(data)

    return new_y

def Catenate(x,y):
    new_list = []
    for i in range(len(x)):
        new_list.append((x[i], y[i]))

    return new_list

def baseline(x, y):
    y = Transpone(y)
    x_y = Catenate(x,y)
    x_y = np.array(x_y)

    v = ConvexHull(x_y).vertices

    v = np.roll(v, -v.argmin())
    v = v[:v.argmax()]

    x = np.array(x)
    y = np.array(y)
    y_baselined = np.interp(x, x[v], y[v])

    return y_baselined


def find_nearest(array, value):
    for data in array:
        if data[0] == value:
            return data[1]

def getData(FileName, Delta):

    wavenumber = []
    transmittance = []

    data = []

    slimmed_wn = []
    slimmed_tr = []

    with open(FileName, "rt", encoding="utf-8") as f:
        
       for line in f:
           if line[0] != "#":
               line = line.split()
               wn = round(float(line[0]),4)
               tr = round(float(line[1]),4)
               
               wavenumber.append(wn)
               transmittance.append(tr)

    for i in range(0, len(wavenumber), Delta):
        slimmed_wn.append(wavenumber[i])
        slimmed_tr.append(transmittance[i])
        data.append((wavenumber[i], transmittance[i]))


    #wavenumber = np.array(slimmed_wn)
    #transmittance = np.array(slimmed_tr)

    wavenumber = slimmed_wn
    transmittance = slimmed_tr

    return wavenumber, transmittance, data


def peakPositions(maximum, minimum):

    peaks = []
    dist_left = float('inf')
    dist_right = float('inf')

    for mini in minimum:
        
        dist_left = float('inf')
        dist_right = float('inf')

        for maxi in maximum:
            if mini > maxi:
                if mini-maxi < dist_left:
                    dist_left = mini - maxi
                    left_max = maxi
            if mini < maxi:
                if maxi-mini < dist_right:
                    dist_right = maxi - mini
                    right_max = maxi
        try:
            peaks.append((float(left_max), float(mini), float(right_max)))
        except:
            pass
    return peaks

def findPeaks(peaks, library):
    foundPeaks = []
    for peak in peaks:
        hits = []
        minimum = peak[0]
        maximum = peak[2]

        for vibration in library.vibrations:
            for data in vibration.frequence:

                maxError = round((abs(maximum - data[0])/data[0])*100,4)
                minError = round((abs(minimum - data[1])/data[1])*100,4)
                sumError = minError + maxError

                if len(hits) < 3:
                    hits.append((data, vibration.name, round(sumError,4), peak))
                else:
                    hits.append((data, vibration.name, round(sumError,4), peak))
                    hits = sorted(hits, key=lambda x: x[2])
                    hits.pop()

        if hits[0][2] <= 5:
            foundPeaks.append((peak, hits))

    return foundPeaks

def contourPeaks(peakPoints, infra):
    
   peakContour = []

   for point in peakPoints:
       peakRange = []
       startPoint = point[0]
       endPoint = point[2]

       for IR in infra:
           if IR[0] >= startPoint and IR[0] <= endPoint:
               peakRange.append(IR)

       peakContour.append(peakRange)
       peakRange = []

   return peakContour


def localExtreme(wavenumber, transmitance):
    
    peaks = find_peaks(transmitance, prominence=1)
    height = transmitance[peaks[0]]
    max_pos = wavenumber[peaks[0]]

    tr_mirrored = transmitance * (-1)
    minima = find_peaks(tr_mirrored, prominence=1)
    min_pos = wavenumber[minima[0]]
    min_heigth = tr_mirrored[minima[0]] * -1

    peaks = peakPositions(max_pos, min_pos)

    return peaks


def drawGraph(IR, wavenumber, transmitance, peakContours, foundPeaks, FileName):

    fig = plt.figure()
    ax = fig.subplots()
    ax.plot(wavenumber, transmitance, color = 'black')
    
    for cont in peakContours:
        xs, ys = zip(*cont)
        ax.fill(xs, ys)


    for i in range(len(foundPeaks)):
        point = foundPeaks[i]
        actX = round(point[0][1],2)
        actY = round(find_nearest(IR, point[0][1]),2)
        ax.scatter(actX, actY, color = "red")
        ax.annotate(i+1, xy=(actX, actY), xytext=(0,-20), 
            textcoords='offset points', ha='center', va='bottom',
            bbox=dict(boxstyle='round,pad=0.2', fc='yellow', alpha=0.3))

    ax.legend()
    ax.grid()
    x_axis = plt.gca()
    x_axis.set_xlim(x_axis.get_xlim()[::-1])
    plt.xlabel("Wavenumber [cm^-1]")
    plt.ylabel("Transmitance [%]")
    plt.savefig("sample.png", dpi=100)
    wb = xl.Workbook()
    ws = wb.worksheets[0]
    img = xl.drawing.image.Image("sample.png")
    img.anchor = 'A1'
    ws.add_image(img)

    ws['N1'] = "Peak"
    ws['O1'] = "Max"
    ws['P1'] = "Min"
    ws['Q1'] = "Type"
    ws['R1'] = "Reference Data"
    ws['S1'] = "Relative error [%]"

    actRow = 2
    for i in range(len(foundPeaks)):
        actPeak = round(foundPeaks[i][0][1], 2)
        actMin = round(foundPeaks[i][0][0], 2)
        actMax = round(foundPeaks[i][0][2], 2)
        ws['N{}'.format(actRow)] = actPeak
        ws['O{}'.format(actRow)] = actMax
        ws['P{}'.format(actRow)] = actMin
        for j in range(len(foundPeaks[i][1])):
            actType = foundPeaks[i][1][j][1]
            actReference = foundPeaks[i][1][j][0]
            actError = foundPeaks[i][1][j][2]
            ws['Q{}'.format(actRow)] = actType
            ws['R{}'.format(actRow)] = "{}-{}".format(actReference[0], actReference[1])
            ws['S{}'.format(actRow)] = actError
            actRow +=1
    autoCellWidth(ws)

    savePath = "{}.xlsx".format(FileName)

    wb.save(savePath)

def newIR(wavenumber, transmitance):

    IR = []
    for i in range(len(wavenumber)):
        IR.append((wavenumber[i], transmitance[i]))
    return IR

def autoCellWidth(worksheet):
    for col in worksheet.columns:
     max_length = 0
     column = col[0].column_letter
     for cell in col:
         try: 
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2
     worksheet.column_dimensions[column].width = adjusted_width

def openFile():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    return file_path

def main(FileName):

    Library = IR_DATA(type='IR_data', wb_name='IR_data.xlsx', sheetname='IR_data')
    wn, tr, sIR = getData("K_3611_4.txt", 1)
    tr = np.array(tr) + baseline(wn, tr)
    wn = np.array(wn)
    peaks = localExtreme(wn, tr)
    foundPeaks = findPeaks(peaks, Library)
    
    hits = []
    for peak in foundPeaks:
        hits.append(peak[0])

    contours = contourPeaks(hits, sIR)
    saveName = FileName.split("/")
    saveName = saveName[-1].split(".")
    saveName = saveName[0]
    drawGraph(sIR, wn, tr, contours, foundPeaks, saveName)

main(openFile())